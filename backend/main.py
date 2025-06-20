from fastapi import FastAPI, HTTPException, UploadFile, File, Form, Query, Depends, Response
from fastapi.middleware.cors import CORSMiddleware
from typing import List, Optional
import mysql.connector
from mysql.connector import Error as MySQLError
import os
import re
import traceback    
import logging
from logging.handlers import RotatingFileHandler
from dotenv import load_dotenv
import subprocess
import json
from datetime import datetime
import bcrypt
from pydantic import BaseModel, Field, field_validator
from fastapi.responses import FileResponse, Response
from openpyxl import Workbook
from io import BytesIO
import zipfile
from contextlib import asynccontextmanager

from gentt import (
    GlobalTimeTableGenerator,
    prepare_timetable_data,
    validate_timetable,
    save_timetables_to_database)

# Load environment variables
load_dotenv()

# Configure logging
def setup_logging():
    logger = logging.getLogger('timetable_api')
    logger.setLevel(logging.INFO)
    
    # File handler with log rotation
    file_handler = RotatingFileHandler(
        'timetable_api.log', 
        maxBytes=10*1024*1024,  # 10MB
        backupCount=5
    )
    # Console handler
    console_handler = logging.StreamHandler()
    
    # Formatter
    formatter = logging.Formatter(
        '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
    
    file_handler.setFormatter(formatter)
    console_handler.setFormatter(formatter)
    
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)
    
    return logger

# Initialize logger
logger = setup_logging()

# Predefined Slots and Days
PREDEFINED_SLOTS = [
    "8:00-8:50", "8:50-9:40", 
    "BREAK",
    "9:50-10:40", "10:40-11:30", 
    "LUNCH",
    "12:20-1:10", "1:10-2:00", "2:00-2:50", "2:50-3:40"
]

PREDEFINED_DAYS = [
    "Monday", "Tuesday", "Wednesday", "Thursday", "Friday"
]

# Add lifespan function before app creation
@asynccontextmanager
async def lifespan(app: FastAPI):
    """
    Lifespan event handler for startup and shutdown events
    """
    # Startup
    logger.info("Starting Timetable Allocation API")
    logger.info(f"Predefined Slots: {PREDEFINED_SLOTS}")
    logger.info(f"Predefined Days: {PREDEFINED_DAYS}")

    uploads_dir = "uploads"
    os.makedirs(uploads_dir, exist_ok=True)

    yield

    # Shutdown
    logger.info("Shutting down Timetable Allocation API")
    try:
        for filename in os.listdir(uploads_dir):
            file_path = os.path.join(uploads_dir, filename)
            try:
                if os.path.isfile(file_path):
                    os.unlink(file_path)
            except Exception as e:
                logger.error(f"Error cleaning up {file_path}: {e}")
    except Exception as e:
        logger.error(f"Error during shutdown cleanup: {e}")

# Create FastAPI app with lifespan
app = FastAPI(
    title="Timetable and Allocator API",
    description="Comprehensive API for timetable generation, validation, and schedule retrieval",
    version="1.0.0",
    lifespan=lifespan
)

# CORS Middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:3000",
        "http://127.0.0.1:3000",
        "https://timetable-backend-tz59.onrender.com",
        "*"
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Pydantic Models
class LoginRequest(BaseModel):
    username: str = Field(..., min_length=1)
    password: str = Field(..., min_length=1)

class ChangePasswordRequest(BaseModel):
    username: str = Field(..., min_length=1)
    oldPassword: str = Field(..., min_length=1)
    newPassword: str = Field(
        ...,
        min_length=8,
        description="Password must contain at least 8 characters, including uppercase, lowercase, numbers, and special characters"
    )

    @field_validator('newPassword')
    @classmethod
    def validate_password(cls, v: str) -> str:
        if len(v) < 8:
            raise ValueError('Password must be at least 8 characters long')
        
        if not any(c.islower() for c in v):
            raise ValueError('Password must contain at least one lowercase letter')
        
        if not any(c.isupper() for c in v):
            raise ValueError('Password must contain at least one uppercase letter')
        
        if not any(c.isdigit() for c in v):
            raise ValueError('Password must contain at least one number')
        
        if not any(c in '@$!%*?&' for c in v):
            raise ValueError('Password must contain at least one special character (@$!%*?&)')
        
        return v

# Database Connection Functions
def get_db_connection():
    """
    Establish database connection
    """
    try:
        conn = mysql.connector.connect(
            host=os.getenv('DB_HOST'),
            user=os.getenv('DB_USER'),
            password=os.getenv('DB_PASSWORD'),
            database=os.getenv('DB_NAME'),
            port=int(os.getenv('DB_PORT'))
        )
        return conn
    except MySQLError as err:
        logger.error(f"Database connection error: {err}")
        raise HTTPException(
            status_code=500, 
            detail=f"Database connection error: {str(err)}"
        )

def get_login_db_connection():
    """
    Establish connection to login database with better error handling
    """
    try:
        # Try to get port from environment, default to 20721 if not set
        port = os.getenv('DB_PORT', '20721')
        
        # Validate port
        try:
            port = int(port)
        except ValueError:
            logger.error(f"Invalid DB_PORT value: {port}")
            port = 20721  # Use default port if invalid
        
        # Get other connection parameters with defaults
        host = os.getenv('DB_HOST', 'localhost')
        user = os.getenv('DB_USER', 'root')
        password = os.getenv('DB_PASSWORD', '')
        
        conn = mysql.connector.connect(
            host=host,
            user=user,
            password=password,
            database='login_details',
            port=port,
            ssl_disabled=False
        )
        return conn
    except MySQLError as err:
        logger.error(f"Login database connection error: {err}")
        raise HTTPException(
            status_code=500,
            detail=f"Database connection error: {str(err)}"
        )
    except Exception as e:
        logger.error(f"Unexpected error in database connection: {e}")
        raise HTTPException(
            status_code=500,
            detail="Internal server error while connecting to database"
        )

# Helper Functions
def verify_password(plain_password: str, hashed_password: str) -> bool:
    """
    Verify password against hashed version
    """
    try:
        return bcrypt.checkpw(
            plain_password.encode('utf-8'),
            hashed_password.encode('utf-8')
        )
    except Exception as e:
        logger.error(f"Password verification error: {e}")
        return False

def hash_password(password: str) -> str:
    """
    Hash password using bcrypt
    """
    try:
        return bcrypt.hashpw(
            password.encode('utf-8'),
            bcrypt.gensalt()
        ).decode('utf-8')
    except Exception as e:
        logger.error(f"Password hashing error: {e}")
        raise HTTPException(
            status_code=500,
            detail="Password hashing failed"
        )

def find_latest_timetable_schema(cursor):
    """
    Find the most recent timetable schema
    """
    try:
        cursor.execute("SHOW DATABASES")
        databases = cursor.fetchall()
        logger.info(f"Available databases: {databases}")

        timetable_schemas = [
            db['Database'] for db in databases 
            if str(db['Database']).startswith('timetable_')
        ]

        if not timetable_schemas:
            raise HTTPException(
                status_code=404, 
                detail="No timetable schema found"
            )

        latest_schema = sorted(timetable_schemas, reverse=True)[0]
        logger.info(f"Selected schema: {latest_schema}")

        return latest_schema
    
    except Exception as e:
        logger.error(f"Error finding timetable schema: {e}")
        raise HTTPException(
            status_code=500, 
            detail=f"Error finding timetable schema: {str(e)}"
        )

def rearrange_timetable_data(timetable_data):
    """
    Rearrange timetable data to match predefined slots and days
    """
    try:
        ordered_timetable = {}
        
        for day in PREDEFINED_DAYS:
            ordered_timetable[day] = {}
            
            for slot in PREDEFINED_SLOTS:
                if slot in ["BREAK", "LUNCH"]:
                    ordered_timetable[day][slot] = slot
                    continue
                
                original_slot_data = timetable_data.get(day, {}).get(slot)
                ordered_timetable[day][slot] = original_slot_data or None

        return ordered_timetable
    except Exception as e:
        logger.error(f"Error rearranging timetable data: {e}")
        raise HTTPException(
            status_code=500, 
            detail=f"Error processing timetable data: {str(e)}"
        )

def safe_json_parse(json_data):
    """
    Safely parse JSON data
    """
    try:
        if not json_data:
            return {}
        
        if isinstance(json_data, str):
            return json.loads(json_data)
        
        return json_data
    except json.JSONDecodeError:
        logger.error(f"JSON decoding error for data: {json_data}")
        return {}

# Authentication Endpoints
@app.post("/api/login")
async def login(login_data: LoginRequest):
    """
    Handle user login
    """
    conn = None
    cursor = None
    try:
        logger.info(f"Login attempt for user: {login_data.username}")
        
        conn = get_login_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute(
            "SELECT * FROM users WHERE username = %s",
            (login_data.username,)
        )
        user = cursor.fetchone()

        if not user:
            logger.warning(f"Login failed: User not found - {login_data.username}")
            raise HTTPException(
                status_code=401,
                detail="Invalid credentials"
            )

        if not verify_password(login_data.password, user['password']):
            logger.warning(f"Login failed: Invalid password - {login_data.username}")
            raise HTTPException(
                status_code=401,
                detail="Invalid credentials"
            )

        logger.info(f"Successful login: {login_data.username}")
        return {
            "success": True,
            "message": "Login successful",
            "username": user['username']
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Login error: {str(e)}\n{traceback.format_exc()}")
        raise HTTPException(
            status_code=500,
            detail="Server error"
        )
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

@app.post("/api/change-password")
async def change_password(change_pwd_data: ChangePasswordRequest):
    """
    Handle password change
    """
    conn = None
    cursor = None
    try:
        logger.info(f"Password change attempt for user: {change_pwd_data.username}")
        
        conn = get_login_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute(
            "SELECT * FROM users WHERE username = %s",
            (change_pwd_data.username,)
        )
        user = cursor.fetchone()

        if not user:
            logger.warning(f"Password change failed: User not found - {change_pwd_data.username}")
            raise HTTPException(
                status_code=401,
                detail="User not found"
            )

        if not verify_password(change_pwd_data.oldPassword, user['password']):
            logger.warning(f"Password change failed: Invalid current password - {change_pwd_data.username}")
            raise HTTPException(
                status_code=401,
                detail="Current password is incorrect"
            )

        hashed_new_password = hash_password(change_pwd_data.newPassword)
        
        cursor.execute(
            "UPDATE users SET password = %s WHERE username = %s",
            (hashed_new_password, change_pwd_data.username)
        )
        conn.commit()

        logger.info(f"Password changed successfully: {change_pwd_data.username}")
        return {
            "success": True,
            "message": "Password updated successfully"
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Password change error: {str(e)}\n{traceback.format_exc()}")
        raise HTTPException(
            status_code=500,
            detail="Server error"
        )
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

# Timetable Management Endpoints
@app.get("/api/timetables/classes")
def get_class_timetables():
    """
    Retrieve class timetables
    """
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        latest_schema = find_latest_timetable_schema(cursor)
        conn.database = latest_schema
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute("""
            SELECT 
                year, 
                section, 
                timetable_data, 
                free_hours, 
                generated_at
            FROM class_timetables
            ORDER BY year, section
        """)
        
        class_timetables = cursor.fetchall()
        
        formatted_timetables = [
            {
                "year": row['year'],
                "section": row['section'],
                "timetable": rearrange_timetable_data(safe_json_parse(row['timetable_data'])),
                "free_hours": safe_json_parse(row['free_hours']),
                "generated_at": str(row['generated_at']) if row['generated_at'] else None
            }
            for row in class_timetables
        ]
        
        return formatted_timetables
    
    except Exception as e:
        logger.error(f"Error retrieving class timetables: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))
    
    finally:
        if conn:
            conn.close()

@app.get("/api/timetables/teachers")
def get_teacher_timetables():
    """
    Retrieve teacher timetables
    """
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        latest_schema = find_latest_timetable_schema(cursor)
        conn.database = latest_schema
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute("""
            SELECT 
                employee_id, 
                teacher_name, 
                timetable_data, 
                free_hours, 
                generated_at
            FROM teacher_timetables
            ORDER BY teacher_name
        """)
        
        teacher_timetables = cursor.fetchall()
        
        formatted_timetables = [
            {
                "employee_id": row['employee_id'],
                "teacher_name": row['teacher_name'],
                "timetable": rearrange_timetable_data(safe_json_parse(row['timetable_data'])),
                "free_hours": safe_json_parse(row['free_hours']),
                "generated_at": str(row['generated_at']) if row['generated_at'] else None
            }
            for row in teacher_timetables
        ]
        
        return formatted_timetables
    
    except Exception as e:
        logger.error(f"Error retrieving teacher timetables: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))
    
    finally:
        if conn:
            conn.close()

@app.get("/api/timetables/venues")
def get_venue_timetables():
    """
    Retrieve venue timetables
    """
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        latest_schema = find_latest_timetable_schema(cursor)
        conn.database = latest_schema
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute("""
            SELECT 
                venue_id, 
                venue_name, 
                timetable_data, 
                free_hours, 
                generated_at
            FROM venue_timetables
            ORDER BY venue_name
        """)
        
        venue_timetables = cursor.fetchall()
        
        formatted_timetables = [
            {
                "venue_id": row['venue_id'],
                "venue_name": row['venue_name'],
                "timetable": rearrange_timetable_data(safe_json_parse(row['timetable_data'])),
                "free_hours": safe_json_parse(row['free_hours']),
                "generated_at": str(row['generated_at']) if row['generated_at'] else None
            }
            for row in venue_timetables
        ]
        
        return formatted_timetables
    
    except Exception as e:
        logger.error(f"Error retrieving venue timetables: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))
    
    finally:
        if conn:
            conn.close()

# File Processing Endpoints
@app.post("/api/process-year-files")
async def process_year_files(
    department_name: str = Form(...),
    files: List[UploadFile] = File(...)
):
    """
    Process year files for a specific department
    """
    try:
        upload_dir = "uploads"
        os.makedirs(upload_dir, exist_ok=True)
        
        file_paths = []
        for file in files:
            import uuid
            unique_filename = f"{uuid.uuid4()}_{file.filename}"
            file_path = os.path.join(upload_dir, unique_filename)
            
            with open(file_path, "wb") as buffer:
                content = await file.read()
                buffer.write(content)
            
            file_paths.append(file_path)
        
        process = subprocess.Popen([
            'python', 
            'process_year_files.py',
            '--schema', department_name,
            '--files', *file_paths
        ], 
            stdout=subprocess.PIPE, 
            stderr=subprocess.PIPE, 
            text=True
        )
        
        stdout, stderr = process.communicate()
        
        for file_path in file_paths:
            try:
                os.unlink(file_path)
            except Exception as e:
                logger.error(f"Error deleting file {file_path}: {e}")
        
        if process.returncode == 0:
            return {
                "success": True,
                "message": "Year files processed successfully",
                "details": {
                    "stdout": stdout,
                    "stderr": stderr
                }
            }
        else:
            return {
                "success": False,
                "error": "Python script execution failed",
                "details": {
                    "stdout": stdout,
                    "stderr": stderr
                },
                "exit_code": process.returncode
            }
    
    except Exception as e:
        logger.error(f"Error processing year files: {str(e)}")
        raise HTTPException(
            status_code=500, 
            detail=f"Server error: {str(e)}"
        )

@app.post("/api/process-faculty-files")
async def process_faculty_files(
    department_name: str = Form(...),
    faculty_list: UploadFile = File(...),
    faculty_preferences: UploadFile = File(...)
):
    """
    Process faculty files for a specific department
    """
    try:
        upload_dir = "uploads"
        os.makedirs(upload_dir, exist_ok=True)
        
        faculty_list_filename = re.sub(r'\W+', '_', faculty_list.filename)
        faculty_pref_filename = re.sub(r'\W+', '_', faculty_preferences.filename)
        
        faculty_list_path = os.path.join(upload_dir, faculty_list_filename)
        faculty_pref_path = os.path.join(upload_dir, faculty_pref_filename)
        
        logger.info(f"Saving faculty list file: {faculty_list_path}")
        logger.info(f"Saving faculty preferences file: {faculty_pref_path}")
        
        with open(faculty_list_path, "wb") as buffer:
            content = await faculty_list.read()
            buffer.write(content)
            logger.info(f"Faculty list file saved. Size: {len(content)} bytes")
        
        with open(faculty_pref_path, "wb") as buffer:
            content = await faculty_preferences.read()
            buffer.write(content)
            logger.info(f"Faculty preferences file saved. Size: {len(content)} bytes")
        
        process = subprocess.Popen([
            'python', 
            'process_faculty_files.py',
            '--schema', department_name,
            '--faculty-list', faculty_list_path,
            '--faculty-pref', faculty_pref_path
        ], 
            stdout=subprocess.PIPE, 
            stderr=subprocess.PIPE, 
            text=True
        )

        try:
            stdout, stderr = process.communicate(timeout=600)  # 10-minute timeout
        except subprocess.TimeoutExpired:
            process.kill()
            stdout, stderr = process.communicate()
            logger.error("Faculty files processing timed out")
            return {
                "success": False,
                "error": "Processing timed out",
                "details": {
                    "stdout": stdout,
                    "stderr": stderr
                }
            }
        
        logger.info(f"STDOUT: {stdout}")
        logger.error(f"STDERR: {stderr}")
        
        try:
            os.unlink(faculty_list_path)
            os.unlink(faculty_pref_path)
        except Exception as cleanup_err:
            logger.error(f"Error during file cleanup: {cleanup_err}")
        
        if process.returncode == 0:
            return {
                "success": True,
                "message": "Faculty files processed successfully",
                "details": {
                    "stdout": stdout,
                    "stderr": stderr
                }
            }
        else:
            return {
                "success": False,
                "error": "Python script execution failed",
                "details": {
                    "stdout": stdout,
                    "stderr": stderr,
                    "exit_code": process.returncode
                }
            }
    
    except Exception as e:
        logger.error(f"Server error during faculty files processing: {e}")
        logger.error(traceback.format_exc())
        
        try:
            if 'faculty_list_path' in locals():
                os.unlink(faculty_list_path)
            if 'faculty_pref_path' in locals():
                os.unlink(faculty_pref_path)
        except Exception as cleanup_err:
            logger.error(f"Error during error cleanup: {cleanup_err}")
        
        raise HTTPException(
            status_code=500, 
            detail=f"Server error: {str(e)}"
        )

@app.get("/api/schemas")
async def get_schemas():
    """
    Retrieve all available database schemas
    """
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute('SHOW DATABASES')
        databases = cursor.fetchall()
        
        logger.info(f"Available databases: {databases}")
        
        system_schemas = [
            'defaultdb', 'information_schema', 
            'mysql', 'performance_schema', 'sys', 'login_details'
        ]
        
        filtered_schemas = [
            db['Database'] for db in databases 
            if db['Database'].lower() not in system_schemas 
            and not db['Database'].lower().startswith('timetable')
        ]
        
        cursor.close()
        conn.close()
        
        return {
            "success": True,
            "schemas": filtered_schemas
        }
    
    except mysql.connector.Error as err:
        logger.error(f"Database error: {err}")
        raise HTTPException(
            status_code=500, 
            detail=f"Database connection error: {str(err)}"
        )
    except Exception as e:
        logger.error(f"Unexpected error: {e}")
        raise HTTPException(
            status_code=500, 
            detail=f"Unexpected error: {str(e)}"
        )

@app.delete("/api/department/{department_name}")
async def delete_department(department_name: str):
    """
    Delete a department (schema) from the database
    """
    conn = None
    cursor = None
    try:
        # Validate department name
        if not department_name:
            raise HTTPException(
                status_code=400,
                detail="Department name is required"
            )
        
        # System schemas that should not be deleted
        protected_schemas = [
            'defaultdb', 'information_schema', 'mysql', 
            'performance_schema', 'sys', 'login_details'
        ]
        
        if department_name.lower() in protected_schemas:
            raise HTTPException(
                status_code=403,
                detail=f"Cannot delete protected schema: {department_name}"
            )
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Check if schema exists
        cursor.execute('SHOW DATABASES')
        databases = cursor.fetchall()
        available_schemas = [db['Database'] for db in databases]
        
        if department_name not in available_schemas:
            raise HTTPException(
                status_code=404,
                detail=f"Department '{department_name}' not found"
            )
        
        # Check if it's a timetable schema (don't delete those from this endpoint)
        if department_name.startswith('timetable_'):
            raise HTTPException(
                status_code=400,
                detail="Use the timetable deletion endpoint for timetable schemas"
            )
        
        # Log the deletion attempt
        logger.info(f"Attempting to delete department: {department_name}")
        
        # Drop the database
        cursor.execute(f"DROP DATABASE `{department_name}`")
        conn.commit()
        
        logger.info(f"Successfully deleted department: {department_name}")
        
        return {
            "success": True,
            "message": f"Department '{department_name}' deleted successfully"
        }
    
    except HTTPException:
        raise
    
    except Exception as e:
        logger.error(f"Error deleting department {department_name}: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to delete department: {str(e)}"
        )
    
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

@app.get("/api/schema/{schema_name}/sortedtable")
async def get_sorted_table(
    schema_name: str,
    limit: Optional[int] = Query(default=100, ge=1, le=1000),
    offset: Optional[int] = Query(default=0, ge=0),
    sort_by: Optional[str] = Query(default=None),
    order: Optional[str] = Query(default='ASC', pattern='^(ASC|DESC)$')  # Changed from regex to pattern
):
    """
    Retrieve SortedTable data
    """
    conn = None
    try:
        # Establish database connection
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Query to get all schemas
        cursor.execute('SHOW DATABASES')
        databases = cursor.fetchall()
        
        # Convert to list of database names
        available_schemas = [
            db['Database'] for db in databases 
            if db['Database'].lower() not in ['information_schema', 'mysql', 'performance_schema', 'sys', 'defaultdb']
        ]
        
        # Check if schema exists (case-insensitive)
        matching_schemas = [
            schema for schema in available_schemas 
            if schema.lower() == schema_name.lower()
        ]
        
        if not matching_schemas:
            raise HTTPException(
                status_code=404, 
                detail=f"Schema '{schema_name}' not found. Available schemas: {available_schemas}",
                headers={"X-Available-Schemas": json.dumps(available_schemas)}
            )
        
        # Use the matched schema (preserving original case)
        actual_schema = matching_schemas[0]
        
        # Switch to the specific database
        conn.database = actual_schema
        cursor = conn.cursor(dictionary=True)
        
        # Check if table exists
        cursor.execute('SHOW TABLES')
        tables = [table['Tables_in_' + actual_schema] for table in cursor.fetchall()]
        
        # Check for possible table variations
        possible_table_names = [
            'SortedTable_SortedTable_xlsx',
            'SortedTable_SortedTable', 
            'sorted_table', 
            'sorted_tables', 
            'sortedtable'
        ]
        
        matching_tables = [
            table for table in tables 
            for possible_name in possible_table_names 
            if table.lower() == possible_name.lower()
        ]
        
        if not matching_tables:
            raise HTTPException(
                status_code=404, 
                detail=f"No sorted table found in schema '{actual_schema}'. Available tables: {tables}",
                headers={"X-Available-Tables": json.dumps(tables)}
            )
        
        # Use the first matching table
        table_name = matching_tables[0]
        
        # Get table columns
        cursor.execute(f'DESCRIBE `{table_name}`')
        columns = cursor.fetchall()
        
        # Determine sort column
        sort_column = sort_by or columns[0]['Field']
        
        # Validate sort column exists
        if not any(col['Field'].lower() == sort_column.lower() for col in columns):
            sort_column = columns[0]['Field']
        
        # Count total rows
        cursor.execute(f'SELECT COUNT(*) as total FROM `{table_name}`')
        total_rows = cursor.fetchone()['total']
        
        # Construct dynamic query
        query = f"""
        SELECT * FROM `{table_name}`
        ORDER BY `{sort_column}` {order}
        LIMIT %s OFFSET %s
        """
        
        cursor.execute(query, (limit, offset))
        rows = cursor.fetchall()
        
        return {
            "success": True,
            "database": actual_schema,
            "tableName": table_name,
            "columns": [
                {
                    "name": col['Field'],
                    "type": col['Type'],
                    "nullable": col['Null'] == 'YES',
                    "key": col['Key']
                } for col in columns
            ],
            "pagination": {
                "total": total_rows,
                "limit": limit,
                "offset": offset,
                "hasMore": offset + len(rows) < total_rows
            },
            "data": rows
        }
    
    except HTTPException as he:
        raise he
    
    except Exception as e:
        logger.error(f"Error retrieving SortedTable: {e}")
        raise HTTPException(
            status_code=500, 
            detail=f"Database error: {str(e)}"
        )
    
    finally:
        if conn:
            conn.close()

@app.get("/api/schema/{schema_name}/sortedtableformatted")
async def get_sorted_table_formatted(
    schema_name: str,
    limit: Optional[int] = Query(default=1000, ge=1, le=5000),
    offset: Optional[int] = Query(default=0, ge=0)
):
    """
    Retrieve SortedTableFormatted data from a specific schema
    """
    conn = None
    try:
        # Establish database connection
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Verify schema exists
        cursor.execute('SHOW DATABASES')
        databases = cursor.fetchall()
        available_schemas = [
            db['Database'] for db in databases 
            if db['Database'].lower() not in ['information_schema', 'mysql', 'performance_schema', 'sys', 'defaultdb']
        ]
        
        matching_schemas = [
            schema for schema in available_schemas 
            if schema.lower() == schema_name.lower()
        ]
        
        if not matching_schemas:
            raise HTTPException(
                status_code=404, 
                detail=f"Schema '{schema_name}' not found."
            )
        
        # Use the matched schema
        actual_schema = matching_schemas[0]
        
        # Switch to the specific database
        conn.database = actual_schema
        cursor = conn.cursor(dictionary=True)
        
        # Check if the SortedTableFormatted table exists (check variations of the name)
        cursor.execute('SHOW TABLES')
        tables = [table['Tables_in_' + actual_schema] for table in cursor.fetchall()]
        
        # Look for variations of SortedTableFormatted
        possible_table_names = [
            'SortedTableFormatted',
            'SortedTable_Formatted',
            'sortedtableformatted',
            'Formatted_SortedTable',
            'SortedTableFormatted_SortedTableFormatted',
            'SortedTableFormatted_SortedTableFormatted_xlsx'
        ]
        
        matching_tables = [
            table for table in tables 
            for possible_name in possible_table_names 
            if table.lower() == possible_name.lower() or possible_name.lower() in table.lower()
        ]
        
        if not matching_tables:
            raise HTTPException(
                status_code=404, 
                detail=f"SortedTableFormatted table not found in schema '{actual_schema}'. Available tables: {tables}"
            )
        
        # Use the first matching table
        table_name = matching_tables[0]
        
        # Count total rows
        cursor.execute(f'SELECT COUNT(*) as total FROM `{table_name}`')
        total_rows = cursor.fetchone()['total']
        
        # Fetch data
        cursor.execute(f'SELECT * FROM `{table_name}` LIMIT %s OFFSET %s', (limit, offset))
        rows = cursor.fetchall()
        
        return {
            "success": True,
            "database": actual_schema,
            "tableName": table_name,
            "pagination": {
                "total": total_rows,
                "limit": limit,
                "offset": offset,
                "hasMore": offset + len(rows) < total_rows
            },
            "data": rows
        }
    
    except HTTPException as he:
        raise he
    
    except Exception as e:
        logger.error(f"Error retrieving SortedTableFormatted: {e}")
        raise HTTPException(
            status_code=500, 
            detail=f"Database error: {str(e)}"
        )
    
    finally:
        if conn:
            conn.close()

@app.get("/api/schema/{schema_name}/uniquesubjects")
async def get_unique_subjects(
    schema_name: str,
    limit: Optional[int] = Query(default=1000, ge=1, le=5000),
    offset: Optional[int] = Query(default=0, ge=0)
):
    """
    Retrieve UniqueSubjects data from a specific schema
    """
    conn = None
    try:
        # Establish database connection
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Verify schema exists
        cursor.execute('SHOW DATABASES')
        databases = cursor.fetchall()
        available_schemas = [
            db['Database'] for db in databases 
            if db['Database'].lower() not in ['information_schema', 'mysql', 'performance_schema', 'sys', 'defaultdb']
        ]
        
        matching_schemas = [
            schema for schema in available_schemas 
            if schema.lower() == schema_name.lower()
        ]
        
        if not matching_schemas:
            raise HTTPException(
                status_code=404, 
                detail=f"Schema '{schema_name}' not found."
            )
        
        # Use the matched schema
        actual_schema = matching_schemas[0]
        
        # Switch to the specific database
        conn.database = actual_schema
        cursor = conn.cursor(dictionary=True)
        
        # Check if the UniqueSubjects table exists (check variations of the name)
        cursor.execute('SHOW TABLES')
        tables = [table['Tables_in_' + actual_schema] for table in cursor.fetchall()]
        
        # Look for variations of UniqueSubjects
        possible_table_names = [
            'UniqueSubjects',
            'Unique_Subjects',
            'uniquesubjects',
            'unique_subjects',
            'UniqueSubjects_UniqueSubjects',
            'UniqueSubjects_UniqueSubjects_xlsx'
        ]
        
        matching_tables = [
            table for table in tables 
            for possible_name in possible_table_names 
            if table.lower() == possible_name.lower() or possible_name.lower() in table.lower()
        ]
        
        if not matching_tables:
            raise HTTPException(
                status_code=404, 
                detail=f"UniqueSubjects table not found in schema '{actual_schema}'. Available tables: {tables}"
            )
        
        # Use the first matching table
        table_name = matching_tables[0]
        
        # Count total rows
        cursor.execute(f'SELECT COUNT(*) as total FROM `{table_name}`')
        total_rows = cursor.fetchone()['total']
        
        # Fetch data
        cursor.execute(f'SELECT * FROM `{table_name}` LIMIT %s OFFSET %s', (limit, offset))
        rows = cursor.fetchall()
        
        return {
            "success": True,
            "database": actual_schema,
            "tableName": table_name,
            "pagination": {
                "total": total_rows,
                "limit": limit,
                "offset": offset,
                "hasMore": offset + len(rows) < total_rows
            },
            "data": rows
        }
    
    except HTTPException as he:
        raise he
    
    except Exception as e:
        logger.error(f"Error retrieving UniqueSubjects: {e}")
        raise HTTPException(
            status_code=500, 
            detail=f"Database error: {str(e)}"
        )
    
    finally:
        if conn:
            conn.close()

@app.get("/health")
def health_check():
    """
    Simple health check endpoint
    """
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SHOW DATABASES')
        databases = cursor.fetchall()
        
        conn.close()
        return {
            "status": "healthy", 
            "database": "connected",
            "current_time": datetime.now().isoformat(),
            "predefined_slots": PREDEFINED_SLOTS,
            "predefined_days": PREDEFINED_DAYS,
            "available_schemas": [
                db[0] for db in databases 
                if not str(db[0]).lower().startswith(('information_schema', 'mysql', 'performance_schema', 'sys', 'login_details'))
            ]
        }
    except Exception as e:
        return {
            "status": "unhealthy", 
            "error": str(e)
        }

# Error Handlers
@app.exception_handler(Exception)
async def global_exception_handler(request, exc):
    """
    Global exception handler for unhandled exceptions
    """
    logger.error(f"Unhandled exception: {exc}")
    return HTTPException(
        status_code=500,
        detail="Internal server error"
    )

# Timetable and Allocator API
@app.get("/api/timetable-schemas")
def get_timetable_schemas():
    """
    Retrieve all available timetable schemas
    """
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute('SHOW DATABASES')
        databases = cursor.fetchall()
        
        # Filter for timetable schemas only
        timetable_schemas = [
            db['Database'] for db in databases 
            if str(db['Database']).startswith('timetable_')
        ]
        
        return {
            "success": True,
            "schemas": timetable_schemas
        }
    
    except Exception as e:
        logger.error(f"Error retrieving timetable schemas: {e}")
        raise HTTPException(
            status_code=500, 
            detail=f"Database error: {str(e)}"
        )
    
    finally:
        if conn:
            conn.close()

@app.delete("/api/timetable-schema/{schema_name}")
def delete_timetable_schema(schema_name: str):
    """
    Delete a specific timetable schema
    """
    conn = None
    try:
        # Validate schema name for safety
        if not schema_name.startswith('timetable_'):
            raise HTTPException(
                status_code=400,
                detail="Invalid timetable schema name. Must start with 'timetable_'"
            )
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Check if schema exists
        cursor.execute('SHOW DATABASES')
        databases = cursor.fetchall()
        
        available_schemas = [db['Database'] for db in databases]
        
        if schema_name not in available_schemas:
            raise HTTPException(
                status_code=404,
                detail=f"Timetable schema '{schema_name}' not found"
            )
        
        # Drop the database
        cursor.execute(f"DROP DATABASE `{schema_name}`")
        
        return {
            "success": True,
            "message": f"Timetable schema '{schema_name}' deleted successfully"
        }
    
    except HTTPException:
        raise
    
    except Exception as e:
        logger.error(f"Error deleting timetable schema: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Database error: {str(e)}"
        )
    
    finally:
        if conn:
            conn.close()

@app.get("/api/timetable/{schema_name}")
def get_specific_timetable(schema_name: str):
    """
    Retrieve timetable data from a specific schema
    """
    conn = None
    try:
        # Validate schema name for safety
        if not schema_name.startswith('timetable_'):
            raise HTTPException(
                status_code=400,
                detail="Invalid timetable schema name. Must start with 'timetable_'"
            )
        
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Check if schema exists
        cursor.execute('SHOW DATABASES')
        databases = cursor.fetchall()
        
        available_schemas = [db['Database'] for db in databases]
        
        if schema_name not in available_schemas:
            raise HTTPException(
                status_code=404,
                detail=f"Timetable schema '{schema_name}' not found"
            )
        
        # Switch to the specific database
        conn.database = schema_name
        cursor = conn.cursor(dictionary=True)
        
        # Get all timetable data
        timetable_data = {
            "classes": [],
            "teachers": [],
            "venues": []
        }
        
        # Get class timetables
        try:
            cursor.execute("""
                SELECT 
                    year, 
                    section, 
                    timetable_data, 
                    free_hours, 
                    generated_at
                FROM class_timetables
                ORDER BY year, section
            """)
            
            class_timetables = cursor.fetchall()
            
            timetable_data["classes"] = [
                {
                    "year": row['year'],
                    "section": row['section'],
                    "timetable": rearrange_timetable_data(safe_json_parse(row['timetable_data'])),
                    "free_hours": safe_json_parse(row['free_hours']),
                    "generated_at": str(row['generated_at']) if row['generated_at'] else None
                }
                for row in class_timetables
            ]
        except Exception as e:
            logger.warning(f"Could not fetch class timetables from {schema_name}: {e}")
        
        # Get teacher timetables
        try:
            cursor.execute("""
                SELECT 
                    employee_id, 
                    teacher_name, 
                    timetable_data, 
                    free_hours, 
                    generated_at
                FROM teacher_timetables
                ORDER BY teacher_name
            """)
            
            teacher_timetables = cursor.fetchall()
            
            timetable_data["teachers"] = [
                {
                    "employee_id": row['employee_id'],
                    "teacher_name": row['teacher_name'],
                    "timetable": rearrange_timetable_data(safe_json_parse(row['timetable_data'])),
                    "free_hours": safe_json_parse(row['free_hours']),
                    "generated_at": str(row['generated_at']) if row['generated_at'] else None
                }
                for row in teacher_timetables
            ]
        except Exception as e:
            logger.warning(f"Could not fetch teacher timetables from {schema_name}: {e}")
        
        # Get venue timetables
        try:
            cursor.execute("""
                SELECT 
                    venue_id, 
                    venue_name, 
                    timetable_data, 
                    free_hours, 
                    generated_at
                FROM venue_timetables
                ORDER BY venue_name
            """)
            
            venue_timetables = cursor.fetchall()
            
            timetable_data["venues"] = [
                {
                    "venue_id": row['venue_id'],
                    "venue_name": row['venue_name'],
                    "timetable": rearrange_timetable_data(safe_json_parse(row['timetable_data'])),
                    "free_hours": safe_json_parse(row['free_hours']),
                    "generated_at": str(row['generated_at']) if row['generated_at'] else None
                }
                for row in venue_timetables
            ]
        except Exception as e:
            logger.warning(f"Could not fetch venue timetables from {schema_name}: {e}")
        
        return {
            "success": True,
            "schema_name": schema_name,
            "timetable_data": timetable_data
        }
    
    except HTTPException:
        raise
    
    except Exception as e:
        logger.error(f"Error retrieving specific timetable: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Database error: {str(e)}"
        )
    
    finally:
        if conn:
            conn.close()

@app.post("/api/generate-timetable")
async def generate_timetable_fastapi(
    sectionConfig: Optional[str] = Form(None),
    faculty: UploadFile = File(...),
    subjects: UploadFile = File(...),
    venues: UploadFile = File(...),
    cdc: UploadFile = File(...),
):
    try:
        # 1. Prepare the data using the function from gentt.py
        files = {
            "faculty": faculty,
            "subjects": subjects,
            "venues": venues,
            "cdc": cdc,
        }
        form = {"sectionConfig": sectionConfig}

        # Read file content before passing to prepare_timetable_data
        files_content = {k: await v.read() for k, v in files.items()}

        # Call the prepare_timetable_data function
        generator, all_sections_data, faculty_df, cdc_df, venues_data = prepare_timetable_data(form, files_content)

        # 2. Generate timetables using the GlobalTimeTableGenerator instance
        logger.info("Starting timetable generation process")
        generation_success = generator.generate_all_timetables(all_sections_data, venues_data)

        if not generation_success:
            logger.error("Failed to generate timetable after multiple attempts")
            raise HTTPException(status_code=500, detail="Failed to generate timetable after multiple attempts")

        # 3. Save timetables to the database
        logger.info("Saving timetables to database")
        connection_uri = os.getenv("DATABASE_URI")
        if not connection_uri:
            raise HTTPException(status_code=500, detail="Database connection URI not configured.")

        save_success = save_timetables_to_database(generator, all_sections_data, faculty_df, cdc_df, venues_data, connection_uri)

        if not save_success:
            logger.error("Timetable generation succeeded but saving to DB failed")
            raise HTTPException(status_code=500, detail="Timetable generation succeeded but saving to DB failed.")

        # 4. Perform validation and return results
        validation_results = validate_timetable(generator, all_sections_data)

        logger.info("Timetable generated and saved successfully")
        return {
            "status": "success",
            "message": "Timetable generated and saved successfully",
            "validation_summary": {
                "subject_hours_valid": validation_results.get("structure_valid", False),
                "venue_clashes": validation_results.get("has_venue_clashes", False)
            }
        }

    except HTTPException as he:
        raise he
    except Exception as e:
        logger.error(f"Exception during timetable generation: {str(e)}\n{traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Timetable generation failed: {str(e)}")

@app.get("/api/timetable/{schema_name}/excel")
async def download_timetables_excel(schema_name: str):
    """
    Generate and download Excel files for timetables
    """
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Check if schema exists
        cursor.execute('SHOW DATABASES')
        databases = cursor.fetchall()
        if schema_name not in [db['Database'] for db in databases]:
            raise HTTPException(status_code=404, detail=f"Schema {schema_name} not found")
        
        # Switch to schema
        conn.database = schema_name
        cursor = conn.cursor(dictionary=True)
        
        # Create Excel files
        excels = {
            'class_timetables.xlsx': create_class_timetables_excel(cursor),
            'teacher_timetables.xlsx': create_teacher_timetables_excel(cursor),
            'venue_timetables.xlsx': create_venue_timetables_excel(cursor)
        }
        
        # Create ZIP file in memory
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for filename, excel_data in excels.items():
                zip_file.writestr(filename, excel_data.getvalue())
        
        # Get the ZIP content
        zip_buffer.seek(0)
        content = zip_buffer.getvalue()
        
        # Create response with proper headers
        return Response(
            content=content,
            media_type='application/zip',
            headers={
                'Content-Disposition': f'attachment; filename=timetables_{schema_name}.zip'
            }
        )
        
    except Exception as e:
        logger.error(f"Error creating Excel files: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if conn:
            conn.close()

def create_class_timetables_excel(cursor):
    wb = Workbook()
    cursor.execute("SELECT * FROM class_timetables ORDER BY year, section")
    class_data = cursor.fetchall()
    
    for item in class_data:
        sheet_name = f"Year{item['year']}-{item['section']}"
        ws = wb.create_sheet(sheet_name)
        
        # Add headers
        ws.append(['Time'] + PREDEFINED_DAYS)
        
        timetable = safe_json_parse(item['timetable_data'])
        
        # Add data rows
        for slot in PREDEFINED_SLOTS:
            row = [slot]
            for day in PREDEFINED_DAYS:
                cell_data = timetable.get(day, {}).get(slot, '')
                if isinstance(cell_data, dict):
                    cell_text = f"{cell_data.get('code', '')}\n{cell_data.get('teacher', '')}"
                    if cell_data.get('venue'):
                        cell_text += f"\n{cell_data['venue']}"
                    row.append(cell_text)
                else:
                    row.append(cell_data or 'FREE')
            ws.append(row)
    
    # Remove default sheet
    del wb['Sheet']
    
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer

def create_teacher_timetables_excel(cursor):
    wb = Workbook()
    cursor.execute("SELECT * FROM teacher_timetables ORDER BY teacher_name")
    teacher_data = cursor.fetchall()
    
    for item in teacher_data:
        sheet_name = item['teacher_name'][:31]  # Excel sheet name length limit
        ws = wb.create_sheet(sheet_name)
        
        # Add headers
        ws.append(['Time'] + PREDEFINED_DAYS)
        
        timetable = safe_json_parse(item['timetable_data'])
        
        # Add data rows
        for slot in PREDEFINED_SLOTS:
            row = [slot]
            for day in PREDEFINED_DAYS:
                cell_data = timetable.get(day, {}).get(slot, '')
                if isinstance(cell_data, dict):
                    cell_text = f"{cell_data.get('code', '')}\n{cell_data.get('year', '')}-{cell_data.get('section', '')}"
                    if cell_data.get('venue'):
                        cell_text += f"\n{cell_data['venue']}"
                    row.append(cell_text)
                else:
                    row.append(cell_data or 'FREE')
            ws.append(row)
    
    # Remove default sheet
    del wb['Sheet']
    
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer

def create_venue_timetables_excel(cursor):
    wb = Workbook()
    cursor.execute("SELECT * FROM venue_timetables ORDER BY venue_name")
    venue_data = cursor.fetchall()
    
    for item in venue_data:
        sheet_name = item['venue_name'][:31]  # Excel sheet name length limit
        ws = wb.create_sheet(sheet_name)
        
        # Add headers
        ws.append(['Time'] + PREDEFINED_DAYS)
        
        timetable = safe_json_parse(item['timetable_data'])
        
        # Add data rows
        for slot in PREDEFINED_SLOTS:
            row = [slot]
            for day in PREDEFINED_DAYS:
                cell_data = timetable.get(day, {}).get(slot, '')
                if isinstance(cell_data, dict):
                    cell_text = f"{cell_data.get('code', '')}\n{cell_data.get('teacher', '')}\n{cell_data.get('year', '')}-{cell_data.get('section', '')}"
                    row.append(cell_text)
                else:
                    row.append(cell_data or 'FREE')
            ws.append(row)
    
    # Remove default sheet
    del wb['Sheet']
    
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)  # Only seek once
    return excel_buffer

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(
        "main:app",
        host="0.0.0.0",
        port=8000,
        reload=True,
        log_level="info"
    )

